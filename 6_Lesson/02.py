import json
import openpyxl


"""This program improves the excel table and converts its data into a .json file."""


def excel_change(file):
    wb = openpyxl.load_workbook(filename=file)

    sheet = wb.worksheets[0]  # Save the worksheet with which we will work in a variable

    col = sheet.max_column  # Get maximum columns
    row = sheet.max_row  # Get maximum rows

    for i in range(1, row + 1):
        sheet.move_range("A{}:ZZ{}".format(i, i), cols=1)  # Add a new column

    for i in range(1, 2):
        sheet.move_range("A{}:ZZ{}".format(i, i), rows=1)  # Add a new row

    for j in range(1, row + 1):  # Here we merge two columns
        sheet['B{}'.format(j)].value = str(sheet.cell(j, 2).value) + ' ' + str(sheet.cell(j, 3).value)
        sheet.merge_cells('B{}:C{}'.format(j, j))

    sheet.merge_cells(  # Here we merge all columns of marks
        '{}:{}'.format(sheet.cell(1, 4).coordinate, sheet.cell(1, col + 1).coordinate)
    )

    for i in range(2, row + 1):  # In this cycle we consider the average mark
        count = 0
        for j in range(4, col + 1):
            if sheet.cell(i, j).value == 'end':
                break
            if sheet.cell(i, j).value is None:
                continue
            else:
                count += 1
        sheet[f'A{i}'] = '=SUM(B{}:Z{})/{}'.format(i, i, count)

    wb.save('New_students.xlsx')


if __name__ == '__main__':
    path_excel = 'Student.xlsx'
    path_json = 'excel_json.json'
    excel_change(path_excel)
